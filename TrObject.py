import datetime
import socket
import json

import openpyxl as openpyxl

'''
1025 - HDD отсутствует (камера)
2049 - HDD OK (регистратор)
1041 - неправильная модель
1537 - HDD отсутствует (регистратор)
1028 - Not connected
0 - нет соединения
4 - connection lost: can't connect
8193 - HDD не отформатирован
513 - ошибка HDD
2065 - неправильная модель
'''





class TrObject():
    """
    Класс объекта трассир
    """

    def __init__(self, name: str, info: dict or None):
        self.name = name
        self.info = info

    @staticmethod
    def loadFromTrassir(SERVER_IP):
        """
        Функция ожидает получения информации с сервера трассир, лочится на sock.listen
        После запуска скрипта на сервере, функция составляет список серверов
        :return: Возвращает словарь со всеми серверами и устройствами
        """
        print("Старт ожидания информации с сервера")
        sock = socket.socket()
        sock.connect((SERVER_IP ,6969))
        print('Соеднинение с сервером:', "office.icstech.ru")
        result = b''
        while True:
            data = sock.recv(4096)
            if not data:
                break
            result += data
        sock.close()
        result = json.loads(result)
        serverList = TrObject.dictToClasses(result["ipDevices"])
        result["scanDateTime"] = str(datetime.datetime.now())
        return result, serverList

    @staticmethod
    def dictToClasses(result):
        """
        Функция принимает выгрузку с клиента трасир в аргумент, а возвращает список объектов класса Сервер
        в объектах Сервер хранится список устройств - объектов Девайс
        :param result:
        :return:
        """
        serverList = []
        # for servName, devices in list(result.items()):
        #     tempS = TrServer(servName, devices)
        #     for device, devInfo in devices.items():
        #         tempD = TrDevice(device, devInfo)
        #         tempS.addDevice(tempD)
        #         del tempD
        #     serverList.append(tempS)
        #     del tempS
        # return serverList

        for servName, devices in list(result.items()):
            tempS = TrServer(servName, devices)
            for device in devices:
                tempD = TrDevice(device["name"], device)
                tempS.addDevice(tempD)
                del tempD
            serverList.append(tempS)
            del tempS
        return serverList
    @staticmethod
    def saveToFile(devices):
        """
        Сохраняет текущее значение в json файл
        """
        with open ("deviceStatus.json", "w") as f:
            json.dump(devices, f, indent=4)

    @staticmethod
    def loadFromFile(devices, path="deviceStatus.json"):
        """
        Заргужает информацию из файла
        """
        with open(path, "r") as f:
            devices = json.load(f)

    @staticmethod
    def modelInfo(info, toExcel=False):
        data = {}
        if toExcel:
            for server, devices in info.items():
                data[server] = {}
                for device in devices:
                    if device["model"] in data[server]:
                        data[server][device["model"]] += 1
                    else:
                        data[server][device["model"]] = 1
            counter = -1
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 13
            try:
                for servName, servDevices in data.items():
                    counter += 2
                    ws.merge_cells(f"A{counter}:C{counter}")
                    ws[f'A{counter}'] = servName
                    ws[f'A{counter}'].fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                         fill_type="solid")
                    counter += 1
                    ws[f'A{counter}'] = "№"
                    ws[f'B{counter}'] = "Наименование"
                    ws[f'C{counter}'] = "Кол-во (шт.)"
                    i = 0
                    for deviceName, deviceCount in servDevices.items():
                        counter += 1
                        i += 1
                        ws[f'A{counter}'] = str(i) + "."
                        ws[f'B{counter}'] = deviceName
                        ws[f'C{counter}'] = deviceCount
            except:
                input("Неудачное преобразование. Файл не соответствует запросу\nEnter для выхода...")
                exit()
            try:
                wb.save("DevicesInfo.xlsx")
            except:
                input("Ошибка сохранения. Возможно таблица открыта. Повторите попытку\nEnter для выхода...")


class TrServer(TrObject):


    def __init__(self, name: str, info: dict or None):
        self.__devices = []
        super().__init__(name, None)

    def addDevice(self,device):
        self.__devices.append(device)

    def offlineDevice(self, *args):
        """
        Показывает оффлайн девайсы на сервере,
        также можно передать аргументы в виде строк, какую информацию надо выносить
        :param args: списко параметров для вывода, если None - выводит объект
        :return:
        """
        result = []
        for device in self.__devices:
            if device.status in [1028, 0, 4]:
                if args == None:
                    result.append(device)
                else:
                    result.append(''.join([f"{device[arg]}  " for arg in args])+'\n')
        return result if len(result) > 0 else None


class TrDevice(TrObject):

    def __init__(self, name: str, info: dict or None):
        super().__init__(name, None)
        self.info = info
        self.status = info["status"] # Int
        self.ip = info["ip"] # string
        self.port = info["port"]
        self.model = info["model"]

    def __getitem__(self, item):
        return self.info[item]





